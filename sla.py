import time as time
import sys as system
import json as javascript_notation
import openpyxl as py_excel
import win32com.client as win32
from bind import Render_Excel
from bind import Contents
from bind import Email_Config
from pst import PST
from conf import Conf
from email import Email
from os import path, listdir
from os import getcwd, chdir, mkdir
from openpyxl import load_workbook
from datetime import datetime
from csv import writer
from treelib import Node, Tree

# HALT CONFIGURATION FILE PATH IF IT IS IN A PROPER 
# FILE PATH FORMAT
def halt_config(file_path: str) -> bool:
    if not path.exists(file_path):        
        system.stdout.write("[ x ] Cannot Find __config__ file for configurations ...")
        return False        
    return True        

# LOAD CONFIGURATION FILE IF IT IS IN A VALID JSON FORMAT
def load_config_file(file_path: str, order: bool) -> dict:
    if not order:
        system.exit("\n... Script Terminated Successfully ...")
    try:
        with open(file=file_path, mode="r", encoding="utf-8") as json_configuration_file:              
            config_file = javascript_notation.loads(json_configuration_file.read())
            json_configuration_file.close()
    except PermissionError as conf_file_permission:        
        system.exit("PermissionError: %s\n... Script Terminated Successfully ..." % conf_file_permission)
    except javascript_notation.decoder.JSONDecodeError as json_error:
        system.exit("JSONDecodeError: %s\n... Script Terminated Successfully ..." % json_error)    
    return config_file

# LOAD WORK BOOK AND VALIDATE FOR ERRORS
def load_excel_work_book(file: str) -> object:
    try:
        work_book = load_workbook(filename=file)            
    except PermissionError as excel_file_permission:
        system.stdout.write("PermissionError: %s" % excel_file_permission)
        return False
    except py_excel.utils.exceptions.InvalidFileException as excel_ext_error:
        system.stdout.write("InvalidFileException: with (%s) %s" % (each_files, excel_ext_error))
        return False
    return work_book

def progressbar(it, prefix="", size=60, file=system.stdout):
    count = len(it)
    def show(j):
        x = int(size*j/count)
        file.write("%s%s%s %i/%i\r" % (prefix, "▐▌"*x, "."*(size-x), j, count))
        file.flush()        
    show(0)
    for i, item in enumerate(it):
        yield item
        show(i+1)
    file.write("\n")
    file.flush()

# ---------------------------------------------------------------
# START OF THE PROGRAM# Email Monitoring and SLA Checker Tool using Python 3.8
# == GLOBAL DATA WAREHOUSE ==
# Project Supervisor: Erickson Flores
# Developed By:
# - Cabrera Troy A.
# - Canolas Jevb John
# - Narido Carlo
# - Palisoc Norman
# ---------------------------------------------------------------

# CONFIGURATION FILE
chdir(path.dirname(path.realpath(__file__)))
configuration_file_path = r"%s\%s" % (getcwd(), r"config\__config__.json")

# # VALIDATE CONFIGURATION FILE
# * CHECK IF FILE EXISTS
# * LOAD FILE IF ITS CONTENTS IS IN A JSON FORMAT
order = halt_config(file_path=configuration_file_path)
config_file = load_config_file(file_path=configuration_file_path, order=order)

# CLASSES INITIALIZATION
conf = Conf(class_pst=PST(), class_email_conf=Email_Config())
content_checker = Contents(email_contents=config_file["Email-Monitoring"]["Email-Contents"])
excel_content = Render_Excel()

# CONFIGURE CONFIG FILE AND GET THE PATH TO SCAN EXCELS
configures = conf.configure(config_file)
path_to_scan = config_file["Email-Monitoring"]["Path-To-Scan-Excels"]

print(open(file=r"config\banner.txt", mode="r", encoding="utf-8").read())
time.sleep(1)

if not path.isdir(path_to_scan):
    print("DirectoryNotFoundError: No '%s' directory to scan, Please check __config__ file." % path_to_scan)
    system.exit("... Script Terminated Successfully ...")
for each_files in listdir(path=path_to_scan):
    if path.splitext(each_files)[1] not in config_file["Email-Monitoring"]["Excel-File-Extension"]:
        continue
    chdir(path_to_scan)
    work_book = load_excel_work_book(file=each_files)
    if work_book == False:
        continue
    logs_array = []
    hierarchy = []
    headers=["Job", "Key Search", "Datetime Received (PST)", "Process", "Average Start/End (Time)", "Average Start/End (Bool)", "SLA (Time)", "SLA (Bool)"]
    logs_array.append(headers) 
    temp_logs_arrays = []
    total_job_counter = 0
    for sheets in work_book.sheetnames:
        work_sheet = work_book[sheets]
        excel_columns = content_checker.yield_excel_columns(work_sheet, work_sheet.max_column)
        if not content_checker.substantiate_email_contents(excel_columns):                                                    
            continue        
        key_count = 0
        for rows in progressbar(range(2, sum([work_sheet.max_row, 1])), "%sChecking Emails: " % (chr(32)*4), 20):
            try:
                valid_columns = config_file["Email-Monitoring"]["Email-Contents"]
                email_address = excel_content.yield_email_address(work_sheet.cell(rows, excel_content.get_column_index(valid_columns[0], excel_columns)).value)
                if email_address is not None:
                    subject = str(work_sheet.cell(rows, excel_content.get_column_index(valid_columns[1], excel_columns)).value)
                    time_received = str(work_sheet.cell(rows, excel_content.get_column_index(valid_columns[2], excel_columns)).value)
                    body = work_sheet.cell(rows, excel_content.get_column_index(valid_columns[3], excel_columns)).value
                    load_end_time = excel_content.yield_load_end_time(body, email_address, config_file)
            except ValueError as value_error:
                # print("ValueError: %s" % value_error)
                continue
            email_contents = dict(email=email_address, subject=subject, time=time_received, load_end_time=load_end_time, body=body)            
            log = conf.collect_valid_email_body(current_datetime=configures["standard_pt"], email_contents=email_contents, valid_emails_list=configures["emails"], configuration_file=config_file)
            if log != None:                
                if log[0] == 'Complete':
                    temp_array = []
                    # print(log[1])
                    dict_logs = conf.validate_job_loads([dict(log[1]), config_file])
                    if dict_logs != None:
                        # print(dict_logs[0])
                        hierarchy.append(dict_logs[0])
                        dict_logs = dict_logs[1]
                        for each_temp in dict_logs:
                            temp_array.append(dict_logs[each_temp])
                        logs_array.append(temp_array)
                        total_job_counter = total_job_counter + 1

            time.sleep(0.00000000001)
    work_book.save(filename=each_files)

    if not path.isdir("Reports"):
        mkdir("Reports")

    with open(r"Reports\R_HIER_L_%s_LOGS.txt" % path.splitext(each_files)[0], "w+", newline='', encoding="utf-8") as hierarchy_logs:
        hierarchy_logs.write(str(conf.hierarchy_structures(hierarchy, "GDWD0000")))
        hierarchy_logs.close()    
    with open(r"Reports\R_CSV_L_%s_LOGS.csv" % path.splitext(each_files)[0], "a+", newline='', encoding="utf-8") as output_logs:
        csv_writer = writer(output_logs)
        csv_writer.writerows(logs_array)
        output_logs.close()    
    with open(r"Reports\LOG_REPORTS_%s_%s.csv" % (path.splitext(each_files)[0], "".join(str(datetime.now()).split(" ")[0].split("-"))), "a+", newline='', encoding="utf-8") as log_reports:
        csv_writer = writer(log_reports)
        csv_writer.writerows([["Email", "Subject", "Time", "Body", "Checked", "In Time"]])
        log_reports.close()

    rows = 2
    total_success = 0
    total_fail = 0
    total_missing = 0
    late_jobs = ""
    missing_jobs = ""

    email_sys = Email(
        recv_email="Troy.Cabrera@email.com",
        report_csv="LOG_REPORTS_%s_%s.csv" % (path.splitext(each_files)[0], "".join(str(datetime.now()).split(" ")[0].split("-"))),
        late_jobs=late_jobs,
        missing_jobs=missing_jobs,
        var_conf=dict(total_success=total_success, total_fail=total_fail, total_missing=total_missing)
    )

    for sheets in work_book.sheetnames:
        if sheets == "Checklist":
            sheet = work_book['Checklist']
            mr = sheet.max_row
            ms_jobs = email_sys.yield_missing_jobs(max_row=mr, sheet=sheet)
            if ms_jobs:
                missing_jobs = ms_jobs["missing_jobs"]
                total_missing = ms_jobs["total_missing"]
        else:
            sheet = work_book[sheets]
            fl_jobs = email_sys.yield_late_jobs(max_row=sheet.max_row, sheet=sheet, rows=rows)
            if fl_jobs:
                late_jobs = fl_jobs["late_jobs"]
                total_fail = fl_jobs["total_fail"]
    
    late_jobs = email_sys.yield_jobs_headers(total_fail=total_fail, total_missing=total_missing)
    email_body = email_sys.construct_email_body(total_job_counter=total_job_counter)
    email_sys.send_email(email_body=email_body, subject="Email Monitoring Report")


