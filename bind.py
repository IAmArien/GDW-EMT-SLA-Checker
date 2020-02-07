import re as regex
import json as javascript_notation
from os import listdir
from os import path
from os import getcwd
from openpyxl import load_workbook
from datetime import datetime

class Render_Excel(object):
    def __init__(self):
        super().__init__()
        self.excel_data = object

    # GET LOAD END TIME FROM THE BODY
    @classmethod
    def yield_load_end_time(self, email_body: str, email_address: str, config_file: dict) -> datetime:
        account = config_file["Email-Monitoring"]["Accounts"][email_address]
        try:
            date = ""
            time = ""
            for each_dt_time_formats in account["Datetime-Tokens"]:
                date = regex.search(account["Datetime-Tokens"][each_dt_time_formats]["Date-Regex"], email_body)
                time = regex.search(account["Datetime-Tokens"][each_dt_time_formats]["Time-Regex"], email_body)                
                if not date == "" and date is not None:
                    date = date.group()
                else:
                    if date == "":
                        date = ""
                if not time == "" and time is not None:
                    time = time.group()
                else:
                    if time == "":
                        time = ""
            self.excel_data = "%s %s" % (date, time)
        except Exception as e:
            print("Regex Error LET: %s" % e)
            return None
        return datetime.strptime(self.excel_data, account["Datetime-Format"])

    # GET EMAIL ADDRESS FROM 'FROM' COLUMN
    @classmethod
    def yield_email_address(self, from_content: str) -> str:
        try:
            email_address = regex.search(r'\S+@\S+', from_content).group()        
            if len(regex.findall('(<|>)', email_address)) > 0 :
                email_address = regex.sub("(<|>)","", email_address)          
            self.excel_data = "%s" % (email_address)
        except (Exception, AttributeError) as e:
            print("Regex Error EA: %s" % e)
            return None
        return self.excel_data

    # GET COLUMN INDEX OF A SPECIFIC EMAIL CONTENT SUCH AS From, To, Subject etc.
    # in a SPECIFIC ROW in an EXCEL FILE
    @staticmethod
    def get_column_index(column_name: str, excel_columns: list) -> int:
        if column_name in excel_columns:
            return sum([excel_columns.index(column_name), 1])
        return 0

# NOTE SYNTAX
# INITIALIZATION
# GETTING LOAD END TIME OF THE EMAIL
# GETTING THE EMAIL ADDRESS OF THE SENDER
# DECLARE THE FOLLOWING SYNTAX BELOW THE CLASS
# --------------------------------------------------------------------------------
# sla_excel = Excel()                                                               // init
# email_address = sla_excel.yield_email_address(work_sheet.cell(rows, cols).value)  // getting 'From' value of an email from a specific rows and cols and retrieve its respective email address
# load_end_time = sla_excel.yield_load_end_time(work_sheet.cell(rows, cols).value)  // getting 'Subject' valud of an email from a specific rows and cols of an excel file and retrieve its respective load end time
# ---------------------------------------------------------------------------------

class Email_Config(object):
    def __init__(self):
        super().__init__()

    # RETRIEVE ALL VALID EMAILS NEEDED
    @staticmethod
    def acquire_valid_emails(email_config_file: str) -> list:
        valid_emails = []
        try:
            with open(r"%s\%s" % (getcwd(), email_config_file), "r", encoding="utf-8") as config_file:
                emails = config_file.read()          
                config_file.close() 
            emails = javascript_notation.loads(emails)       
            for each_email in emails["Email-Monitoring"]["Accounts"]:
                if len(regex.findall("@x*", each_email)) > 0:          
                    valid_emails.append(each_email) 
        except FileNotFoundError as file_error:
            print("Config File Not Found: %s" % file_error)
            return None       
        return valid_emails    

    # GET THE DIRECTORY AND FILENAME OF config FILE CONTAINING THE VALID EMAILS
    @staticmethod
    def get_email_config_file(directory: str, config_extension: str) -> str:        
        if path.isdir(r"%s\%s" % (getcwd(), directory)):
            for each_files in listdir(path=directory):
                spl_filename = path.splitext(each_files)                     
                if spl_filename[1] == config_extension:
                    return r"%s\{filename}{extension}".format(filename=spl_filename[0], extension=spl_filename[1]) % directory
            else:
                return None
        else:
            print("Path '{directory}' not found".format(directory=directory))
            return None

# NOTE SYNTAX
# INITIALIZATION
# GETTING directory of config file containing the list of valid emails
# OBTAINING the list of valid emails from the directory given
# DECLARE THE FOLLOWING CODES BELOW THE CLASS
# --------------------------------------------------------------
# >>> config = Email_Config()                                   // init
# >>> config_file = config.get_email_config_file("emails")      // get directory
# >>> config.acquire_valid_emails(config_file)                  // get valid emails
# --------------------------------------------------------------

class Contents(object):
    def __init__(self, email_contents: list):
        super().__init__()
        self.email_contents = email_contents    
    
    # CHECK FOR THE REQUIRED VALID EMAILS
    def substantiate_email_contents(self, email_columns: list) -> bool:
        valid_columns_count = []
        for each_email_columns in email_columns:
            if each_email_columns in self.email_contents:
                valid_columns_count.append(each_email_columns)        
        return len(valid_columns_count) == len(self.email_contents)

    # GET EXCEL COLUMNS IN AN EXCEL FILE
    def yield_excel_columns(self, work_sheet: object, columns: int) -> list:
        excel_columns = []
        for i in range(1, sum([columns, 1])):
            excel_columns.append(work_sheet.cell(1, i).value)
        return excel_columns

# NOTE SYNTAX
# INITIALIZATION
# CHECK ONLY FOR THE REQUIRED VALID EMAIL COLUMNS FROM THE __config__ FILE
# GET THE EXCEL COLUMNS GIVEN IN AN EXCEL FILE
# DECLARE THE FOLLOWING CODES BELOW THE CLASS
# --------------------------------------------------------------------------------------------------
# >>> content_checker = Contents(email_contents=config_file["Email-Monitoring"]["Email-Contents"])  // init
# >>> excel_columns = content_checker.yield_excel_columns(work_sheet, work_sheet.max_column)        // check columns from an excel file
# >>> content_checker.substantiate_email_contents(excel_columns)                                    // validate columns with available valid columns from the __config__ file
# --------------------------------------------------------------------------------------------------


